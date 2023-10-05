# %%
# script using the eniscope API to retrive organization alarm settings and control monitored equipment work our of working hours
# version 1.1
import eniscopeapi as es
import pandas as pd
import credentials as cr
import time, datetime
import os, ast, pprint

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

# %%

config_file = "orgs_to_monitor.cfg"
UPLOAD_FILES = True

# define list of organizations and equipment datachannels to be monitored
default_config = {
    "Burger King Vallecas": [
        "PLAYK",
        "CLIMA LOBBY",
        "CLIMA KITCHEN",
        "BROILER",
        "EXTRACTOR (fryers and broiler)",
        "COFFEE MAKER",
        "WARM TABLE",
        "SEVERAL LIGHTS / KITCHEN&LOCKERS&WAREHOUSE LIGHTS / SALOON LIGHTS",
        "LIGHTS_SALOON&WC",
        "BEACONS",
        "CHIPS&FRITS_FREEZERS",
    ],
    "Burger King Benalmádena": [
        "KITCHEN CLIMA",
        "LIGHTING LOBBY 1",
        "FRYER 1",
        "FRYER 2",
        "FRYER 3",
        "FRYER 4",
        "KITCHEN EXTRACTOR",
        "BROILER",
        "FRITS-CHIPS FRIDGE & MICROVAWE",
        "WARM TABLE",
        "SANDWICHER",
        "LOBBY CLIMA",
    ],
}
# check if configuration file orgs_to_monitor.cfg exist, if yes, then read it into monitoring list dictionary, if not - then create a new one and store as JSON default_monitoring_list


config_file = "orgs_to_monitor.conf"


def write_default_config():
    with open(config_file, "w") as f:
        f.write(pprint.pformat(default_config))


def current_time():
    return datetime.datetime.now().strftime("[%Y-%m-%d %H:%M:%S]: ")


def read_config():
    with open(config_file, "r") as f:
        content = f.read()
        try:
            return ast.literal_eval(content)
        except ValueError:
            print(
                f"{current_time()}Error: Configuration file is not a valid Python dictionary."
            )
            return {}


# Check if the configuration file exists
if not os.path.exists(config_file):
    write_default_config()
    monitoring_list = default_config
else:
    monitoring_list = read_config()

    # report export to excel file
folder_path = f"./reports"
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# check if hwminutes_summary.csv exists, then read it to dataframe hwminutes_summary
if not os.path.exists("./reports/hwminutes_summary.xlsx"):
    hwminutes_summary = pd.DataFrame()
else:
    hwminutes_summary = pd.read_excel("./reports/hwminutes_summary.xlsx")


# %%
start_time = time.time()
# create the API object
api = es.EniscopeAPIClient(cr.api_key)

# authenticate the API object
if not api.authenticate_user():
    print(f"{current_time()}Authentication failed")
    exit()
else:
    print(f"{current_time()}Authentication successful")

# Run dtata collection and report prepare for ech organisation in the monitoring list

for org_to_monitor in monitoring_list.keys():
    # get organization id for the organization to be monitored
    print(
        f'{current_time()}Getting organization id for "{org_to_monitor}"...',
        end="",
        flush=True,
    )
    org = api.get_organizations_list(organization_name=org_to_monitor)[0]
    print("done")
    org_id = org["organizationId"]

    # get list of channels for the organization
    print(
        f'{current_time()}Getting list of channels for "{org_to_monitor}"...',
        end="",
        flush=True,
    )
    channels = api.get_channels_list(organization_id=org_id)
    print("done")

    # retrive the alarm settings for the organization
    print(
        f'{current_time()}Getting alarm settings for "{org_to_monitor}"...',
        end="",
        flush=True,
    )
    alarms, rules, periods = api.get_alarm_data(organization_id=org_id)
    print("done")

    # %%
    # create dataframes for channels, alarms, rules and periods for the whole organization

    print(f"{current_time()}Alarms data prerocessing...", end="")

    channels_df = pd.DataFrame.from_dict(channels)
    alarms_df = pd.DataFrame.from_dict(alarms)
    rules_df = pd.DataFrame.from_dict(rules)
    periods_df = pd.DataFrame.from_dict(periods)
    alarms_df["channelName"] = alarms_df["channelId"].apply(
        lambda x: channels_df.loc[
            channels_df["dataChannelId"] == x, "channelName"
        ].iloc[0]
    )

    channels_columns_to_int = ["dataChannelId"]
    alarms_colums_to_int = [
        "alarmId",
        "channelId",
        "organizationId",
        "reportingInterval",
        "reminderInterval",
        "status",
    ]
    rules_columns_to_int = ["alarmRuleId", "alarmId"]
    periods_columns_to_int = ["alarmPeriodId", "alarmId"]

    channels_df[channels_columns_to_int] = channels_df[channels_columns_to_int].apply(
        lambda x: pd.to_numeric(x, errors="ignore", downcast="integer")
    )
    alarms_df[alarms_colums_to_int] = alarms_df[alarms_colums_to_int].apply(
        lambda x: pd.to_numeric(x, errors="ignore", downcast="integer")
    )
    rules_df[rules_columns_to_int] = rules_df[rules_columns_to_int].apply(
        lambda x: pd.to_numeric(x, errors="ignore", downcast="integer")
    )
    rules_df["thresholdValue"] = rules_df["thresholdValue"].apply(
        lambda x: pd.to_numeric(x, errors="ignore", downcast="float")
    )
    periods_df[periods_columns_to_int] = periods_df[periods_columns_to_int].apply(
        lambda x: pd.to_numeric(x, errors="ignore", downcast="integer")
    )

    # %%

    # convert period_df days column from string to list of integers
    periods_df["days"] = periods_df["days"].apply(
        lambda x: [int(i) for i in x.split(",")]
    )

    # Merge alarms_df, rules_df, and periods_df on 'alarmId'
    merged_df = pd.merge(alarms_df, rules_df, on="alarmId", how="inner")
    merged_df = pd.merge(merged_df, periods_df, on="alarmId", how="inner")

    # Select the columns you want to include in the final result
    selected_columns = [
        "alarmId",
        "alarmName",
        "channelId",
        "channelName",
        "emailRecipients",
        "emailTemplateId",
        "emailLanguage",
        "alarmInterval",
        "reportingInterval",
        "reminderInterval",
        "status",
        "expires",
        "timeZone",
        "alarmRuleId",
        "field",
        "thresholdType",
        "thresholdDirection",
        "thresholdValue",
        "thresholdPeriod",
        "alarmPeriodId",
        "days",
        "startTime",
        "endTime",
        "startDate",
        "endDate",
    ]

    # Extract the selected columns and transpose to create a one-row DataFrame
    alarm_settings = merged_df[selected_columns]

    # filter out dataframes for the channels, alarms, rules and periods to be monitored

    alarms_to_monitor = alarm_settings[
        alarm_settings["channelName"].isin(monitoring_list[org_to_monitor])
    ]

    print("done")

    # %%
    # Portion of code to pull last day of data for monitored channels
    # set the start and end dates for the data pull. Integer Unix time normalized to midnight and linked to the Organization timezone
    startTimestamp = (
        int(
            pd.to_datetime("now", utc=True)
            .tz_convert(org["timeZone"])
            .normalize()
            .timestamp()
        )
        - 86400
    )
    endTimestamp = int(
        pd.to_datetime("now", utc=True)
        .tz_convert(org["timeZone"])
        .normalize()
        .timestamp()
    )
    fields = list(alarms_to_monitor["field"].unique())
    # add Energy meter into a list if thereis not E in the list
    if "E" not in fields:
        fields.append("E")

    print(
        f'{current_time()}Geting channels readings for {org_to_monitor} for {pd.to_datetime(startTimestamp, unit="s", utc=True).tz_convert(org["timeZone"]).date()}...',
        end="",
        flush=True,
    )
    channel_data = api.get_multiple_channel_data(
        list(alarms_to_monitor["channelId"].unique()),
        [(startTimestamp, endTimestamp)],
        fields=fields,
    )
    channel_data_df = pd.DataFrame()
    for channel in channel_data.values():
        df = pd.DataFrame(channel["records"])
        df["channelId"] = channel["channel"]
        df["channelName"] = channel["name"]
        df["datetime"] = pd.to_datetime(df["ts"], unit="s", utc=True).dt.tz_convert(
            org["timeZone"]
        )
        channel_data_df = pd.concat([channel_data_df, df], ignore_index=True)

    channel_data_df

    print("done")

    # %%

    import eniscopedata as ed

    report = pd.DataFrame()

    print(f"{current_time()}Calculating alarms activation...", end="", flush=True)

    for channel_id in sorted(channel_data_df["channelId"].unique().tolist()):
        ch_alarms = alarms_to_monitor[alarms_to_monitor["channelId"] == channel_id]
        for alarm in ch_alarms.iterrows():
            alarm = alarm[1]
            if alarm.status == 1:
                rule = ed.Threshold(
                    alarm.thresholdValue,
                    alarm.thresholdDirection,
                    alarm.field,
                    alarm.reportingInterval,
                )
                schedule = ed.Schedule(
                    set(alarm.days),
                    [alarm.startTime, alarm.endTime],
                    tz=org["timeZone"],
                )

                # checkif startDate is valid
                if alarm.startDate != None:
                    startDate = (
                        pd.to_datetime(alarm.startDate, utc=True)
                        .tz_convert(org["timeZone"])
                        .timestamp()
                    )
                else:
                    startDate = pd.to_datetime("1970/01/01").timestamp()
                # checkif endDate is valid
                if alarm.endDate != None:
                    endDate = (
                        pd.to_datetime(alarm.endDate, utc=True)
                        .tz_convert(org["timeZone"])
                        .timestamp()
                    )
                else:
                    endDate = pd.to_datetime("2038/01/19").timestamp()

                # Calculate mean value for the 'field' column for the period defined by reportingInterval, NaN data filled with closed future value. Calculation stored in a new column named {field}_mean.
                channel_data_df.loc[
                    channel_data_df["channelId"] == channel_id, f"{rule.field}_mean"
                ] = (
                    channel_data_df.loc[
                        channel_data_df["channelId"] == channel_id, rule.field
                    ]
                    .rolling(rule.reportInterval)
                    .mean()
                    .bfill()
                )
                # check is alarm is active in certain time in accordance to the schedule of alarms and thresholds breaks
                alarm_active_1 = channel_data_df["channelId"] == channel_id
                alarm_active_2 = schedule == channel_data_df["datetime"]
                alarm_active_3 = rule == channel_data_df[f"{rule.field}_mean"]
                alarm_active = alarm_active_1 & alarm_active_2 & alarm_active_3
                is_active = (alarm_active).sum()
                # add a new column to the dataframe named {field}_alarm_active with the result of the previous check
                if f"{rule.field}_alarm_active" not in channel_data_df.columns:
                    channel_data_df[f"{rule.field}_alarm_active"] = False
                channel_data_df.loc[alarm_active, f"{rule.field}_alarm_active"] = bool(
                    True
                )
                # add resulsts to the report
                if is_active != 0:
                    report = pd.concat(
                        [
                            report,
                            ed.createReport(
                                org_to_monitor,
                                channel_data_df,
                                channel_id,
                                alarm.alarmName,
                                schedule,
                                rule,
                                is_active,
                            ),
                        ],
                        ignore_index=True,
                    )
                print(".", end="", flush=True)
    print("done")

    # %%
    # code to calculate summary for the report columns Active Time, HH:mm	and Energy consumed, kWh. Active time is a string  in hh:mm format and need to be converted to timedelta to be able to sum it
    # define funtion which takes a string in hh:mm format and returns a timedelta object and ass it to a timedelta sum
    totalTime = pd.Timedelta(0)

    def activeTime(x):
        a_timedelta = pd.Timedelta(f"{x}:00")
        global totalTime
        totalTime += a_timedelta

    report["Active Time, HH:mm"].apply(activeTime)
    totalTime = str(totalTime)[-8:-3]
    report.sort_values("Energy consumed, kWh", ascending=False, inplace=True)

    report_sum = pd.concat(
        [
            report,
            pd.DataFrame(
                [
                    {
                        "Organization": "",
                        "Equipment": "SUMMARY",
                        "Alarm": "",
                        "Alarm Rule": "",
                        "Schedule": "",
                        "Active Time, HH:mm": totalTime,
                        "Energy consumed, kWh": report["Energy consumed, kWh"].sum(),
                    }
                ]
            ),
        ]
    )

    # %%

    # Define the Excel file path
    file_path = f"./reports/{org_to_monitor}_alarms_report.xlsx"

    reportDate = (
        pd.to_datetime(startTimestamp, unit="s", utc=True)
        .tz_convert(org["timeZone"])
        .strftime("%Y-%m-%d")
    )

    # # Check if the Excel file exists
    # if os.path.exists(file_path):
    #     # If it exists, load it
    #     with pd.ExcelFile(file_path, engine="openpyxl") as xls:
    #         if f"Report_{reportDate}" in xls.sheet_names:
    #             existing_df = pd.read_excel(xls, f"Report_{reportDate}")
    #         else:
    #             existing_df = None

    # Write (or overwrite) the specific sheet with openpyxl
    # Determine the mode for the ExcelWriter ('a' for append if file exists, 'w' otherwise)
    write_mode = "a" if os.path.exists(file_path) else "w"
    with pd.ExcelWriter(file_path, engine="openpyxl", mode=write_mode) as writer:
        if f"Report_{reportDate}" in writer.book.sheetnames:
            del writer.book[f"Report_{reportDate}"]
        report_sum.to_excel(writer, index=False, sheet_name=f"Report_{reportDate}")
        workbook = writer.book
        worksheet = writer.sheets[f"Report_{reportDate}"]

        # Formatting
        wrap_alignment = Alignment(wrap_text=True)
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Header format
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        header_format = {
            "font": header_font,
            "fill": header_fill,
            "alignment": header_alignment,
        }
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Set column widths and formats
        for col in worksheet["A:C"]:
            col[0].alignment = wrap_alignment
            worksheet.column_dimensions[col[0].column_letter].width = 20

        for col in worksheet["D:G"]:
            col[0].alignment = center_alignment
            worksheet.column_dimensions[col[0].column_letter].width = 15

        # insert date column to report_sum and remove summary line
        report_sum.insert(0, "Date", reportDate)
        report_sum = report_sum.iloc[:-1]
    print(f"{current_time()}Report for {org_to_monitor} is ready.\n")

    # check is lines with same date and organization exist in hwminutes_summary dataframe and delete them, then append report_sum to hwminutes_summary
    if not hwminutes_summary.empty:
        hwminutes_summary = hwminutes_summary[
            ~(
                (hwminutes_summary["Date"] == reportDate)
                & (hwminutes_summary["Organization"] == org_to_monitor)
            )
        ]
        hwminutes_summary = pd.concat(
            [hwminutes_summary, report_sum], ignore_index=True
        )
    else:
        hwminutes_summary = report_sum.copy()

    print(f"{current_time()}Summary for {org_to_monitor} is updated.\n")

# save updated hwminutes_summary to csv file
hwminutes_summary.to_excel("./reports/hwminutes_summary.xlsx", index=False)


# %%
if UPLOAD_FILES == True:
    from oauth2client.service_account import ServiceAccountCredentials
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    from googleapiclient.http import MediaFileUpload
    from google.oauth2.service_account import Credentials as ServiceAccountCredentials
    import os

    # function which upload files to a given google drive folder

    print("{current_time()}Copy reports and updated summary to a Google Drive:")

    def get_file_id(drive_service, folder_id, file_name):
        """Get the file ID of a file in a specific folder by its name."""
        query = f"'{folder_id}' in parents and name='{file_name}'"
        results = (
            drive_service.files().list(q=query, fields="files(id, name)").execute()
        )
        files = results.get("files", [])

        if not files:
            return None
        return files[0]["id"]

    def upload_to_drive(drive_service, folder_id, file_path):
        """
        Upload a file to a given folder on Google Drive.
        Parameters:
        - drive_service: The Drive API service instance.
        - folder_id: The ID of the folder to upload the file to.
        - file_path: The path to the files to upload.
        """
        file_name = os.path.basename(file_path)
        file_metadata = {"name": file_name, "parents": [folder_id]}

        # Check if the file already exists in the folder
        existing_file_id = get_file_id(drive_service, folder_id, file_name)

        media = MediaFileUpload(file_path, resumable=True)

        try:
            if existing_file_id:
                # Update the existing file (Note: We remove the 'parents' key from the metadata)
                update_metadata = {"name": file_name}
                request = drive_service.files().update(
                    fileId=existing_file_id,
                    body=update_metadata,
                    media_body=media,
                    fields="id",
                )
                file_info = request.execute()
                print(
                    f"\t{current_time()}Updated {file_path} on Drive, File ID: {file_info['id']}"
                )
            else:
                # Create a new file
                request = drive_service.files().create(
                    body=file_metadata, media_body=media, fields="id"
                )
                file_info = request.execute()
                print(
                    f"\t{current_time()}Uploaded {file_path} to Drive, File ID: {file_info['id']}"
                )
        except HttpError as error:
            print(f"\t{current_time()}An error occurred: {error}")

    # Routine that perfom Authentication and upload steps
    # Path to the service account JSON key file

    service_account_file = "feedbackloop-399807-300aec3efe37.json"

    # The ID of the folder where you want to upload the file.
    # You can get this from the folder's URL on Google Drive: https://drive.google.com/drive/folders/YOUR_FOLDER_ID
    FOLDER_ID = "1G1YUlZZQV52sBZ1lpefHcV2EbPb8u8M-"  # the folder ID of Projects/HW Minutes/ folder to store the files

    creds = ServiceAccountCredentials.from_service_account_file(
        service_account_file, scopes=["https://www.googleapis.com/auth/drive.file"]
    )

    # Build the Drive API client once
    drive_service = build("drive", "v3", credentials=creds)

    # Path to the folder containing the files you want to upload
    folder_path = "./reports"

    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)

        # Check if it's a regular file (and not a directory)
        if os.path.isfile(file_path):
            upload_to_drive(drive_service, FOLDER_ID, file_path)

print(
    f"\n{current_time()}Total reports prepare time: {time.time() - start_time} seconds.\n{current_time()}All done."
)
